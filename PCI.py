
import os
import openpyxl as opxl
import pdfplumber as pdfp
import win32com.client as win32
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def GetPDFPASS(path, filename):
    pdf = pdfp.open(path + "\\" + filename +".pdf") # open the pdf

    ''''''''''''''''''''''''''''''''''''''''''
    ''' if pdf format change, modify here  '''
    ''''''''''''''''''''''''''''''''''''''''''
    # set parameters:
    # page means which page has the result(pass/fail)
    # key means which represent pass
    # line menas which line has the result
    if path.find("CLK") > -1:
        page = 2
        key = "(Pass)"
        line = 1
    elif path.find("Gen4") > -1 :
        page = 0
        key = "Pass!"
        line = 4
    elif path.find("Gen5") > -1:
        page = 0
        key = "PASS"
        line = 3

    text = pdf.pages[page].extract_text().split('\n') # get each line text on specific page's  
    checkpass = text[line].find(key) # get the result(pass or fail) poistion in result line  
    ''''''''''''''''''''''''''''''''''''''''''
    '''                end                 '''
    ''''''''''''''''''''''''''''''''''''''''''

    
    if checkpass > -1:
        return "Pass"
    else :
        return "Fail"

def GetTXTPASS(path, filename):
    with open(path + "\\" + filename +".txt", 'r', encoding='utf-8') as file:
        file.readline() # introduction line
        for content in file.readlines(): # get each line's context
            content = content.split('\t') # split the context by tab

            ''''''''''''''''''''''''''''''''''''''''''
            ''' if txt format change, modify here  '''
            ''''''''''''''''''''''''''''''''''''''''''
            # 12 means each line is consist of 12 sections
            # 2 means "Pass" in the third section(0, 1, 2)
            if ( len(content) == 12 and content[2] != 'Pass'): 
                return "Fail"
            ''''''''''''''''''''''''''''''''''''''''''
            '''                end                 '''
            ''''''''''''''''''''''''''''''''''''''''''  

        return "Pass"

def ExcelFormat(wb):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    if "Sheet" in wb.sheetnames:
        wb["Sheet"].title = "PCI_Express"

    ws = wb["PCI_Express"]
    ws.column_dimensions['A'].width = 3.25
    ws.column_dimensions['B'].width = 20.11
    ws.column_dimensions['C'].width = 21.67
    ws.column_dimensions['D'].width = 21.67
    ws.column_dimensions['E'].width = 14.11
    ws.column_dimensions['F'].width = 21.67
    ws.column_dimensions['G'].width = 14.11
    ws.column_dimensions['H'].width = 21.67
    ws.column_dimensions['I'].width = 14.11
    ws.column_dimensions['J'].width = 21.67
    ws.column_dimensions['K'].width = 12.89
    ws.column_dimensions['L'].width = 21.6
    ws.column_dimensions['M'].width = 12.89
    ws.column_dimensions['N'].width = 12.89
    ws.column_dimensions['O'].width = 12.89
    ws.column_dimensions['P'].width = 12.89
    ws.column_dimensions['Q'].width = 12.89
    ws.column_dimensions['R'].width = 16.22
    ws.column_dimensions['S'].width = 10
    ws.column_dimensions['T'].width = 10
    ws.column_dimensions['U'].width = 10
    ws.column_dimensions['V'].width = 10

    # 設定標題
    ws['A1'] = "Unit Test: PCI Express"
    ws['A1'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # 合併單元格
    ws.merge_cells('A1:V1')

    ws['C2'] = "Tests Planned"
    ws['C3'] = "Tests Blocked"
    ws['C4'] = "Tests Passed"
    ws['C5'] = "Tests Failed"
    
    ws['A2'] = f'=SUM(S35:S2225)/2' 
    ws['A3'] = f'=SUM(T35:T2225)/2' 
    ws['A4'] = f'=SUM(U35:U2225)/2' 
    ws['A5'] = f'=SUM(V35:V2225)/2' 

    ws.merge_cells('A2:B2')
    ws.merge_cells('A3:B3')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A5:B5')

    ws['A2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")

    ws['C2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws['A2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['B2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['B3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['B4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['B5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A5'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws.merge_cells('C2:G2')
    ws.merge_cells('C3:G3')
    ws.merge_cells('C4:G4')
    ws.merge_cells('C5:G5')

    ###### UUT PART ########
    ws['A7'] = "UUT Information"
    ws['A7'].font = Font(name='Tahoma', size=12, bold=True, color="000000")
    ws['A7'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A8'] = "Project"
    ws['A8'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A9'] = "Project Revision"
    ws['A9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A10'] = "Part Number"
    ws['A10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A11'] = "Series Number"
    ws['A11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A12'] = "PEX89144 BIOS Version"
    ws['A12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A13'] = "COM Express"
    ws['A13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A14'] = "DIMM"
    ws['A14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A15'] = "Cable CEM"
    ws['A15'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A16'] = "Power Supply"
    ws['A16'].font = Font(name='Tahoma', size=10, color="000000")

    # let user input the UUT information 
    ws['D8'] = input("enter Project: ")
    ws['D8'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D9'] = input("enter Project Revision: ")
    ws['D9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D10'] = input("enter Part Number: ")
    ws['D10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D11'] = input("enter Series Number: ")
    ws['D11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D12'] = input("enter PEX89144 BIOS Version: ")
    ws['D12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D13'] = input("enter COM Express: ")
    ws['D13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D14'] = input("enter DIMM: ")
    ws['D14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D15'] = input("enter Cable CEM: ")
    ws['D15'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D16'] = input("enter Power Supply: ")
    ws['D16'].font = Font(name='Tahoma', size=10, color="000000")   
    # end of UUT input
   
    ws[f'A{7}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'A{8}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{9}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{10}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{11}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{12}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{13}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{14}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{15}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{16}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{7}'].border = Border(top=medium, left=thin, right=medium, bottom=thin)
    ws[f'D{8}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{9}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{10}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{11}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{12}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{13}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{14}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{15}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{16}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws.merge_cells('A7:G7')
    ws.merge_cells('A8:C8')
    ws.merge_cells('A9:C9')
    ws.merge_cells('A10:C10')
    ws.merge_cells('A11:C11')
    ws.merge_cells('A12:C12')
    ws.merge_cells('A13:C13')
    ws.merge_cells('A14:C14')
    ws.merge_cells('A15:C15')
    ws.merge_cells('A16:C16')

    ws.merge_cells('D8:G8')
    ws.merge_cells('D9:G9')
    ws.merge_cells('D10:G10')
    ws.merge_cells('D11:G11')
    ws.merge_cells('D12:G12')
    ws.merge_cells('D13:G13')
    ws.merge_cells('D14:G14')
    ws.merge_cells('D15:G15')
    ws.merge_cells('D16:G16')

    ##### TEST EQUIPMENT PART ###############
    ws['A19'] = "Test Equipment:"
    ws['A19'].font = Font(name='Tahoma', size=10, bold=True)
    ws.merge_cells('A19:B19')

    ws['A20'] = "1."
    ws['A20'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A21'] = "2."
    ws['A21'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A22'] = "3."
    ws['A22'].font = Font(name='Tahoma', size=10, color="000000")

    ''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Equipment change '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B20'] = "Tektronix DPO75002SX ATI Performance Oscilloscope 50GHz / 200GS/s"
    ws['B20'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B21'] = "Tektronix P7625 TriMode Probe Bandwidth 25G *2"
    ws['B21'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B22'] = "PCIe 5.0 32 GT/s CEM Electrical Test Fixture(CLB, CBB, ISI)"
    ws['B22'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''


    ###### Test Auxiliary Application PART ##############
    ws['A25'] = "TEST AUXILIARY APPLICATION"
    ws['A25'].font = Font(name='Tahoma', size=10, bold=True)
    ws.merge_cells('A25:C25')

    ws['A26'] = "1."
    ws['A26'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A27'] = "2."
    ws['A27'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A28'] = "3."
    ws['A28'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A29'] = "4."
    ws['A29'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A30'] = "5."
    ws['A30'].font = Font(name='Tahoma', size=10, color="000000")

    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Auxiliary Application change '''
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ws['B26'] = "Tektronix DPO75002SX Version 10.8.7 Build 29"
    ws['B26'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B27'] = "Intel ClockJitter Tool  5.0.2"
    ws['B27'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B28'] = "Intel SigTest 3.2.0.3"
    ws['B28'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B29'] = "Intel SigTest 4.0.52"
    ws['B29'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B30'] = "Intel SigTest Phoenix 5.1.04"
    ws['B30'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws.merge_cells('B26:D26')
    ws.merge_cells('B27:C27')

    ##### TEST PROCEDURE - TX PART ###########
    ws['A33'] = "Test Procedure - TX"
    ws['A33'].font = Font(name='Tahoma', size=10, bold=True)
    ws.merge_cells('A33:B33')

    ws['A34'] = "1."
    ws['A34'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A35'] = "2."
    ws['A35'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A36'] = "3."
    ws['A36'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A37'] = "4."
    ws['A37'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A38'] = "5."
    ws['A38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A39'] = "6."
    ws['A39'].font = Font(name='Tahoma', size=10, color="000000")

    ''''''''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if Test Procedure - TX change '''
    ''''''''''''''''''''''''''''''''''''''''''''''''''
    ws['B34'] = "Warm up Scope 20 mins at least. Calibrate and deskew the SMA Cables. "
    ws['B34'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B35'] = "Connect CLB/CBB with UUT."
    ws['B35'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B36'] = "Connec two SMA cables between CLB/CBB and Scope."
    ws['B36'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B37'] = "Capture the waveform and transfer to .wfm file on Scope."
    ws['B37'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B38'] = "Run SigTest software to validate TX Signal Quality & Preset Test."
    ws['B38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B39'] = "Run ClockJitter software to validate 100M Reference Clock(RefCLK) Test"
    ws['B39'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ###### REF DOCUMENT PART #############
    ws['A42'] = "Ref Document"
    ws['A42'].font = Font(name='Tahoma', size=10, bold=True)
    ws.merge_cells('A42:B42')

    ws['A43'] = "1."
    ws['A43'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A44'] = "2."
    ws['A43'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A45'] = "3."
    ws['A45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A46'] = "4."
    ws['A46'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A47'] = "5."
    ws['A47'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A48'] = "6."
    ws['A48'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A49'] = "7."
    ws['A49'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A50'] = "8."
    ws['A50'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A51'] = "9."
    ws['A51'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A52'] = "10."
    ws['A52'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A53'] = "11."
    ws['A53'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if REF DOCUMENT change  '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B43'] = "PCI Express 3.0 Signal Quality testing for Systems using Tektronix MSO/DPO/DSA70K Series Real Time Oscilloscopes Version 1.0."
    ws['B43'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B44'] = "PCI Express Base Specification Revision 2.1, March 4, 2009."
    ws['B44'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B45'] = "PCI Express Card Electromechanical Specification Revision 2.0, April 11, 2007."
    ws['B45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B46'] = "PCI Express Base Specification Revision 3.0 Version 1.0, November 10, 2010."
    ws['B46'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B47'] = "PCI Express Card Electromechanical Specification Revision 3.0, July 21, 2013."
    ws['B47'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B48'] = "PCI Express Base Specification Revision 4.0 Version 1.0, September 27, 2017."
    ws['B48'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B49'] = "PCI Express Card Electromechanical Specification Revision 4.0, Version 0.9, November 27, 2018."
    ws['B49'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B50'] = "PCI Express® Base Specification Revision 5.0 Version 1.0, May 22, 2019."
    ws['B50'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B51'] = "PCI Express Card Electromechanical Specification Revision 5.0, Version 1.0, June 9, 2021."
    ws['B51'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B52'] = "Tektronix PCIE Gen4 TX CEM MOI, Version 0.7, Feb-2019."
    ws['B52'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B53'] = "646758_PCIe* 5.0 Compliance Test BKM Revision 1.0, August 2021."
    ws['B53'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ws.merge_cells('A56:V57')
    ws['A56'] = "PCIE Slot - Tx Signal Quality"
    ws['A56'].font = Font(name='Tahoma', size=16, bold=True, color="FFFFFF")
    ws['A56'].fill = PatternFill(start_color="1E81B0", end_color="1E81B0", fill_type="solid")

#######################################

def Create_New_Excel():
    newwb = opxl.Workbook() # create a new excel with one default sheet, sheet name is "sheet"
    newwb.save("Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx") # name the excel "Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx"

def Get_Path():
    root = input("enter file path: ex.D:\SW1tool\QT_SI_Report_Auto\BP-PEX003-BD_Rev.A01\n") # user input the root file path
    # root = "D:\\SW1tool\\QT_SI_Report_Auto\\BP-PEX003-BD_Rev.A01"
    path = root + "\\Result_SI\\PCIE_Tx"
    return path, root

def PCIEJGENZ_Sort(name):
    # custom_sort
    match = re.search(r'(\D+)(\d+)', name) # get the number in filename
    if match:
        prefix = match.group(1)  # 'PCIE' or 'JGENZ'
        number = int(match.group(2))  # The numeric part
        return (prefix != 'PCIE', number) # PCIE first, then JGENZ
    
    return (True, 0)  # no match the form, be the last one

def Preset(temppath):
    global g_plan, g_block, g_pass, g_fail
    preset4path = temppath+"\\Preset\Gen4_Preset"
    preset4list = []
    preset4file = []
    presetdata = [0, 0, 0, 0] # planned, blocked, passed, failed

    try:
        presetdata[0] += 1
        gen4_preset = GetTXTPASS(preset4path, "preset_results") # try to get the result
    except: # blocked!!! 
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", preset4path + "\\preset_results.txt")

        preset4list.append("preset_results not found") 
        preset4file.append(None)
        presetdata[1] += 1
    else:
        preset4list.append(gen4_preset) # store the result 
        preset4file.append(preset4path + "\\preset_results.txt") # store the file path
        if gen4_preset == "Pass": 
            presetdata[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("fail test result: ", preset4path + "\\preset_results.txt")
            presetdata[3] += 1

    preset5path = temppath+"\\Preset\Gen5_Preset"
    preset5list = []
    preset5file = []
    os.chdir(preset5path)
    try:
        presetdata[0] += 1
        gen5_preset = GetPDFPASS(preset5path, "PCIe Gen 5 Preset Test AC Results")
    except:
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", preset5path + "\\PCIe Gen 5 Preset Test AC Results.pdf")

        preset5list.append("PCIe Gen 5 Preset Test AC Results not found")
        preset5file.append(None)
        presetdata[1] += 1
    else:
        preset5list.append(gen5_preset)
        preset5file.append(preset5path + "\\PCIe Gen 5 Preset Test AC Results.pdf")
        if gen5_preset == "Pass":
            presetdata[2] += 1
        else:
            print("!!!!!!!!!!!!!")
            print("!!! Block !!!")
            print("!!!!!!!!!!!!!")
            print("fail test result: ", preset5path + "\\PCIe Gen 5 Preset Test AC Results.pdf")
            presetdata[3] += 1

    return preset4file, preset5file, preset4list, preset5list, presetdata
    
def Clk(temppath):
    global g_plan, g_block, g_pass, g_fail
    clkpath = temppath + "\\CLK_Jitter" 
    clklist = []
    clkfile = []
    clkdata = [0, 0, 0, 0] # planned, blocked, passed, failed
    for i in range (1, 6):
        try:
            clkdata[0] += 1
            clk = GetPDFPASS(clkpath, "PCIE" + str(i) + ".0")
        except: # block!!!
            print("!!!!!!!!!!!!!")
            print("!!! Block !!!")
            print("!!!!!!!!!!!!!")
            print("miss file: ", clkpath + "\\PCIE" + str(i) + ".0.pdf")
            clklist.append("PCIE" + str(i) + ".0 not found")
            clkfile.append(None)
            clkdata[1] += 1
        else:
            clklist.append(clk)
            clkfile.append(str(clkpath) + "\\PCIE" + str(i) + ".0.pdf")
            if clk == "Pass":
                clkdata[2] += 1
            else:
                print("!!!!!!!!!!!!!!")
                print("!!!  Fail  !!!")
                print("!!!!!!!!!!!!!!")
                print("Fail test result: ", clkpath + "\\PCIE" + str(i) + ".0.pdf")
                clkdata[3] += 1

    return clkfile, clklist, clkdata

def SignalQT(temppath, i):
    global g_plan, g_block, g_pass, g_fail
    gen4path = temppath + "\\Gen4"
    gen4list = []
    gen4file = []
    gendata = [0, 0, 0, 0] # planned, blocked, passed, failed
    try:
        gendata[0] += 1
        gen4 = GetPDFPASS(gen4path, "Lane" + str(i))
    except: # block!!!
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", gen4path + "\\Lane" + str(i) + ".pdf")
        gen4list.append("Lane" + str(i) + ".0 not found")
        gen4file.append(None)
        gendata[1] += 1
    else:
        gen4list.append(gen4)
        gen4file.append(str(gen4path) + "\\Lane" + str(i) + ".pdf")
        if gen4 == "Pass":
            gendata[2] += 1 
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", gen4path + "\\Lane" + str(i) + ".pdf")
            gendata[3] += 1

    gen5path = temppath + "\\Gen5"
    gen5list = []
    gen5file = []
    try:
        gendata[0] += 1
        gen5 = GetPDFPASS(gen5path, "Lane" + str(i))
    except:
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", gen5path + "\\Lane" + str(i) + ".pdf")
        gen5list.append("Lane" + str(i) + " not found")
        gen5file.append(None)
        gendata[1] += 1
    else:
        gen5list.append(gen5)
        gen5file.append(str(gen5path) + "\\Lane" + str(i) + ".pdf")
        if gen5 == "Pass":
            gendata[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", gen5path + "\\Lane" + str(i) + ".pdf")
            gendata[3] += 1

    return gen4file, gen5file, gen4list, gen5list, gendata

def TestResult(column, row, filepath, result, ws, dict):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'{chr(ord(column)+1)}{row}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    # column, row mean where i want to fill in. ex. C=A, R=1 => A1


    start = str(column) + str(row)
    end = str(column) + str(row+1)
    ws.merge_cells(f'{start}:{end}')
    if filepath != None:
        dict[start] = filepath
    else:
        ws[start].value = "No Result file"
        ws[start].font = Font(name='Tahoma', size=10, color="FF0000")

    ###### font  stysle #####
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws.row_dimensions[row].height=27
    #########################
    
    start = chr(ord(column)+1) + str(row)
    end = chr(ord(column)+1) + str(row+1)
    ws.merge_cells(f'{start}:{end}')
    if result == "Fail":
        ws[start].value = result
        ###### font  stysle #####
        ws[start].font = Font(name='Tahoma', size=10, color="FF0000")
        ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        #########################
    elif result == "Pass":
        ws[start].value = result
        ###### font  stysle #####
        ws[start].font = Font(name='Tahoma', size=10, color="000000")
        ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        #########################   
    else:
        ws[start].value = "no result"
        ###### font  stysle #####
        ws[start].font = Font(name='Tahoma', size=10, color="FF0000")
        ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ######################### 
   
def TestData(column, row, data, ws):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'{chr(ord(column)+1)}{row}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'{chr(ord(column)+2)}{row}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'{chr(ord(column)+3)}{row}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    start = chr(ord(column)) + str(row)
    end = chr(ord(column)) + str(row+1)
    ws.merge_cells(f'{start}:{end}')
    ws[start].value = data[0]
    ws[start].font = Font(name='Tahoma', size=10, color="000000")
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


    start = chr(ord(column)+1) + str(row)
    end = chr(ord(column)+1) + str(row+1)
    ws.merge_cells(f'{start}:{end}')    
    ws[start].value = data[1]
    ws[start].font = Font(name='Tahoma', size=10, color="000000")
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


    start = chr(ord(column)+2) + str(row)
    end = chr(ord(column)+2) + str(row+1)
    ws.merge_cells(f'{start}:{end}')    
    ws[start].value = data[2]
    ws[start].font = Font(name='Tahoma', size=10, color="000000")
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    

    start = chr(ord(column)+3) + str(row)
    end = chr(ord(column)+3) + str(row+1)
    ws.merge_cells(f'{start}:{end}')
    ws[start].value = data[3]
    ws[start].font = Font(name='Tahoma', size=10, color="000000")
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

def Tx_Signal_Quality_xxxxx_x16_Slot_Header(ws, row, name):
    ws.row_dimensions[row].height = 15
    ws.merge_cells(f'A{row}:V{row}')

    if 'PCIE' in name:
        name = 'J'+ name

    ws[f'A{row}'] = "Tx Signal Quality - " + name + "x16 Slot"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")   
    ws[f'A{row}'].alignment = Alignment(vertical="center")

def Tx_Preset_Test_Header(ws, row):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)
    ws.merge_cells(f'A{row}:V{row}')
    ws.row_dimensions[row + 1].height = 14
    ws.row_dimensions[row + 2].height = 60

    ws[f'A{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'D{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)

    ws[f'S{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'S{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'T{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'U{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'V{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws.merge_cells(f'A{row + 1}:B{row + 2}')
    ws.merge_cells(f'C{row + 1}:C{row + 2}')
    ws.merge_cells(f'D{row + 1}:D{row + 2}')
    ws.merge_cells(f'E{row + 1}:E{row + 2}') 
    ws.merge_cells(f'F{row + 1}:F{row + 2}')
    ws.merge_cells(f'G{row + 1}:G{row + 2}')
    ws.merge_cells(f'S{row + 1}:V{row + 1}')
    ws.merge_cells(f'H{row + 1}:R{row + 2}')

    ws[f'A{row}'] = "Tx Preset Test"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(vertical='bottom')

    ws[f'A{row + 1}'] = "Signal Name"
    ws[f'A{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'A{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'C{row + 1}'] = "Comment"
    ws[f'C{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'C{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'D{row + 1}'] = "Gen4\nPreset Test\nResults"
    ws[f'D{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'D{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'E{row + 1}'] = "Preset Test\nPass/Fail"
    ws[f'E{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'E{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'F{row + 1}'] = "Gen5\nPreset Test\nResults"
    ws[f'F{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'F{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'G{row + 1}'] = "Preset Test\nPass/Fail"
    ws[f'G{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'G{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True) 

    ws[f'H{row + 1}'] = "Notes"
    ws[f'H{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'H{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'S{row + 1}'].value = "Number of Tests"
    ws[f'S{row + 1}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'S{row + 2}'].value = "Planned"
    ws[f'S{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'T{row + 2}'].value = "Blocked"
    ws[f'T{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'T{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'T{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'U{row + 2}'].value = "Passed"
    ws[f'U{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'U{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'U{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'V{row + 2}'].value = "Failed"
    ws[f'V{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'V{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'V{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

def Tx_100M_Clock_Jitter_Test_Header(ws, row):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'D{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'I{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'J{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'K{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'L{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'M{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 1}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)

    ws[f'S{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'S{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'T{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'U{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'V{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws.merge_cells(f'A{row}:V{row}')

    ws[f'A{row}'] = "Tx 100M Clock Jitter Test"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(vertical='bottom')
    
    ws.row_dimensions[row + 1].height = 14.1
    ws.row_dimensions[row + 2].height = 60.2

    ws.merge_cells(f'A{row + 1}:B{row + 2}')
    ws.merge_cells(f'C{row + 1}:C{row + 2}')
    ws.merge_cells(f'D{row + 1}:D{row + 2}')
    ws.merge_cells(f'E{row + 1}:E{row + 2}')
    ws.merge_cells(f'F{row + 1}:F{row + 2}')
    ws.merge_cells(f'G{row + 1}:G{row + 2}')
    ws.merge_cells(f'H{row + 1}:H{row + 2}')
    ws.merge_cells(f'I{row + 1}:I{row + 2}')
    ws.merge_cells(f'J{row + 1}:J{row + 2}')
    ws.merge_cells(f'K{row + 1}:K{row + 2}')
    ws.merge_cells(f'L{row + 1}:L{row + 2}')
    ws.merge_cells(f'M{row + 1}:M{row + 2}')
    ws.merge_cells(f'S{row + 1}:V{row + 1}')
    ws.merge_cells(f'N{row + 1}:R{row + 2}')

    ws[f'A{row + 1}'] = "Template File"
    ws[f'A{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'A{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'C{row + 1}'] = "Comment"
    ws[f'C{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'C{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'D{row + 1}'] = "Gen1\nTest Results"
    ws[f'D{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'D{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'E{row + 1}'] = "Gen1 Test\nPass/Fail"
    ws[f'E{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'E{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'F{row + 1}'] = "Gen2\nTest Results"
    ws[f'F{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'F{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'G{row + 1}'] = "Gen2 Test\nPass/Fail"
    ws[f'G{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'G{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'H{row + 1}'] = "Gen3\nTest Results"
    ws[f'H{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'H{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'H{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'I{row + 1}'] = "Gen3 Test\nPass/Fail"
    ws[f'I{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'I{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'J{row + 1}'] = "Gen4\nTest Results"
    ws[f'J{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'J{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'J{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'K{row + 1}'] = "Gen4 Test\nPass/Fail"
    ws[f'K{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'K{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'L{row + 1}'] = "Gen5\nTest Results"
    ws[f'L{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'L{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'L{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 1}'] = "Gen5 Test\nPass/Fail"
    ws[f'M{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'M{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws.merge_cells(f'N{row + 1}:R{row + 2}')
    ws[f'N{row + 1}'] = "Notes"
    ws[f'N{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'N{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'S{row + 1}'].value = "Number of Tests"
    ws[f'S{row + 1}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'S{row + 2}'].value = "Planned"
    ws[f'S{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'T{row + 2}'].value = "Blocked"
    ws[f'T{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'T{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'T{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'U{row + 2}'].value = "Passed"
    ws[f'U{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'U{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'U{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'V{row + 2}'].value = "Failed"
    ws[f'V{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'V{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'V{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

def Tx_Signal_Quality_Test_Header(ws, row):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'D{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)

    ws[f'S{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'S{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'T{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'U{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'V{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws.merge_cells(f'A{row}:V{row}')

    ws[f'A{row}'] = "Tx Signal Quality Test"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(vertical='bottom')
    
    ws.row_dimensions[row + 1].height = 14.1
    ws.row_dimensions[row + 2].height = 60.2

    ws.merge_cells(f'A{row + 1}:B{row + 2}')
    ws.merge_cells(f'C{row + 1}:C{row + 2}')
    ws.merge_cells(f'D{row + 1}:D{row + 2}')
    ws.merge_cells(f'E{row + 1}:E{row + 2}') 
    ws.merge_cells(f'F{row + 1}:F{row + 2}')
    ws.merge_cells(f'G{row + 1}:G{row + 2}')
    ws.merge_cells(f'S{row + 1}:V{row + 1}')
    ws.merge_cells(f'H{row + 1}:R{row + 2}')

    ws[f'A{row + 1}'] = "Signal Name"
    ws[f'A{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'A{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'C{row + 1}'] = "Comment"
    ws[f'C{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'C{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws[f'D{row + 1}'] = "Gen4\nTest Results"
    ws[f'D{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'D{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'E{row + 1}'] = "Gen4\nPass/Fail"
    ws[f'E{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'E{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'F{row + 1}'] = "Gen5\nTest Results"
    ws[f'F{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'F{row + 1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'G{row + 1}'] = "Gen5\nPass/Fail"
    ws[f'G{row + 1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'G{row + 1}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True) 

    ws[f'S{row + 1}'].value = "Number of Tests"
    ws[f'S{row + 1}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'S{row + 2}'].value = "Planned"
    ws[f'S{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'S{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'S{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'T{row + 2}'].value = "Blocked"
    ws[f'T{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'T{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'T{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'U{row + 2}'].value = "Passed"
    ws[f'U{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'U{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'U{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'V{row + 2}'].value = "Failed"
    ws[f'V{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'V{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'V{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws.merge_cells(f'H{row + 1}:R{row + 2}')
    ws[f'H{row + 1}'] = "Notes"
    ws[f'H{row + 1}'].font = Font(name='Tahoma', size=10, bold=True )
    ws[f'H{row + 1}'].alignment = Alignment(horizontal='center', vertical='bottom')

def SignalName(ws, name, row, num):
    if 'PCIE' in name:
        name = 'J'+ name
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    # 調整列寬
    ws.row_dimensions[row].height = 27
    ws.row_dimensions[row + 1].height = 27

    ws[f'A{row}'] = name + "_TXP" + str(num)
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row+1}'] = name + "_TXP" + str(num)
    ws[f'A{row+1}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'A{row+1}'].alignment = Alignment(horizontal='left', vertical='center')

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=thin, bottom=thin)
    ws[f'B{row}'].border = Border(top=medium, left=medium, right=thin, bottom=thin)
    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'B{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)

def Template_File(ws, name, row):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=thin, bottom=thin)
    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'B{row}'].border = Border(top=medium, left=medium, right=thin, bottom=thin)
    ws[f'B{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)

    if 'PCIE' in name:
        name = 'J'+ name
    # 調整列寬
    ws.row_dimensions[row].height = 27
    ws.row_dimensions[row + 1].height = 27

    ws[f'A{row}'] = name + "_RWFCLK+"
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row + 1}'] = name + "_RWFCLK-"
    ws[f'A{row + 1}'].font = Font(name='Tahoma', size=10, color="000000") 
    ws[f'A{row + 1}'].alignment = Alignment(horizontal='left', vertical='center')

def Notes(ws, row, column):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)
    # 調整列寬
    ws.row_dimensions[row].height = 27
    ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")

    ws.merge_cells(f'{column}{row}:R{row + 1}')

def Comment(ws, row):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'C{row}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    # 調整列寬
    ws.row_dimensions[row].height = 27
    ws.merge_cells(f'C{row}:C{row + 1}')
    ws[f'C{row}'] = "Compliance Load board"
    ws[f'C{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')

def NumOfTests(row1, row2, ws):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'S{row2+1}'].border = Border(top=medium, left=medium, right=thin, bottom=medium)
    ws[f'T{row2+1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'U{row2+1}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'V{row2+1}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)

    ws[f'R{row2+1}'] = 'Totals'
    ws[f'R{row2+1}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'R{row2+1}'].alignment = Alignment(horizontal="right", vertical="center", wrapText=True)
    
    ws[f'S{row2+1}'].value = f'=SUM(S{row1}:S{row2})'
    ws[f'S{row2+1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'S{row2+1}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'T{row2+1}'].value = f'=SUM(T{row1}:T{row2})'
    ws[f'T{row2+1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'T{row2+1}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'U{row2+1}'].value = f'=SUM(U{row1}:U{row2})'
    ws[f'U{row2+1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'U{row2+1}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'V{row2+1}'].value = f'=SUM(V{row1}:V{row2})'
    ws[f'V{row2+1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'V{row2+1}'].font = Font(name='Tahoma', size=10, bold =True)

def Tx_Preset_Test_Result(row, name, file4path, f4result, file5path, f5result, ws, data, num, dicts):
    SignalName(ws, name, row, num)
    Comment(ws, row)
    TestResult('D', row, file4path[0], f4result[0], ws, dicts)
    TestResult('F', row, file5path[0], f5result[0], ws, dicts)
    Notes(ws, row, 'H')
    TestData('S', row, data, ws)
    NumOfTests(row, row+1, ws)

def Tx_100M_Clock_Jitter_Test_Result(row, name, filepath, fresult, ws, data, dicts):
    Template_File(ws, name, row)
    Comment(ws, row)
    TestResult('D', row, filepath[0], fresult[0], ws, dicts)
    TestResult('F', row, filepath[1], fresult[1], ws, dicts)
    TestResult('H', row, filepath[2], fresult[2], ws, dicts)
    TestResult('J', row, filepath[3], fresult[3], ws, dicts)
    TestResult('L', row, filepath[4], fresult[4], ws, dicts)
    Notes(ws, row, 'N')
    TestData('S', row, data, ws)
    NumOfTests(row, row+1, ws)

def Tx_Signal_Quality_Test_Result(row, name, gen4file, gen5file, gen4list, gen5list, ws, data, x, dicts):
    SignalName(ws, name, row, x)
    Comment(ws, row)
    TestResult('D', row, gen4file[0], gen4list[0], ws, dicts)
    TestResult('F', row, gen5file[0], gen5list[0], ws, dicts)
    Notes(ws, row, 'H')
    TestData('S', row, data, ws)
    if x == 15:
        NumOfTests(row-30, row+1, ws)

def Insertfile(dicts, root):
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    xl.Visible = False
    xl.ScreenUpdating = False
    xl.DisplayAlerts = False
    wb = xl.Workbooks.Open(f'{root}\\Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx')
    ws = wb.Worksheets("PCI_Express")

    for key in dicts:
        print("insert:", dicts[key])
        dest_cell = ws.Range(key)
        filename = dicts[key].split('\\')
        obj = ws.OLEObjects()
        obj.Add(Filename=dicts[key], Link=False, Left=dest_cell.Left, Top=dest_cell.Top, DisplayAsIcon=True, IconIndex=0, IconLabel=filename[-1], IconFileName="")
        obj.ShapeRange.LockAspectRatio = False
        obj.Height=53
        obj.Width=88
        print("===============================================================================================================")
    
    print("saving...")
    wb.Save()
    xl.Application.Quit()

def main():
    path, root = Get_Path()
    os.chdir(root)
    Create_New_Excel()
    dirs = os.listdir(path)
    dirs.sort(key=PCIEJGENZ_Sort)
    wb = opxl.load_workbook(f'{root}\\Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx')
    ws = wb.active
    ExcelFormat(wb)

    dicts = {}
    for i, dir in enumerate(dirs):
        print("----------------------", dir, "--------------------------------")
        temppath = path + "\\" + dir
        Tx_Signal_Quality_xxxxx_x16_Slot_Header(ws, 59 + i * 59, str(dir))
        Tx_Preset_Test_Header(ws, 61+i*59)
        Tx_100M_Clock_Jitter_Test_Header(ws, 70+i*59)
        Tx_Signal_Quality_Test_Header(ws, 79+i*59)
        preset4file, preset5file, preset4list, preset5list, presetdata = Preset(temppath)
        Tx_Preset_Test_Result(64+i*59, dir, preset4file, preset4list, preset5file, preset5list, ws, presetdata, 0, dicts)

        clkfile, clklist, clkdata = Clk(temppath)
        Tx_100M_Clock_Jitter_Test_Result(73+i*59, dir, clkfile, clklist, ws, clkdata, dicts)

        for x in range(0,16):
            gen4file, gen5file, gen4list, gen5list, gendata = SignalQT(temppath, x)
            Tx_Signal_Quality_Test_Result(82+x*2+i*59, dir, gen4file, gen5file, gen4list, gen5list, ws, gendata, x, dicts)

    wb.save(f'{root}\\Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx')

    Insertfile(dicts, root)

    print("============ DONE ============")

if __name__ == '__main__':
    main()
