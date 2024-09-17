import os
import openpyxl as opxl
import pdfplumber as pdfp
import win32com.client as win32
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def GetPDFPASS(file):
    pdf = pdfp.open(file) # open the pdf

    ''''''''''''''''''''''''''''''''''''''''''
    ''' if pdf format change, modify here  '''
    ''''''''''''''''''''''''''''''''''''''''''
    # set parameters:
    # page means which page has the result(pass/fail)
    # key means which represent pass
    # line menas which line has the result
    if file.find("Droop") > -1 :
        page = 0
        key = "Pass"
        line = 5
    elif file.find("High_Speed_") > -1:
        page = 0
        key = "Pass"
        line = 5
    elif file.find("High_Speed") > -1:
        page = 0
        key = "Pass*"
        line = 5        

    text = pdf.pages[page].extract_text().split('\n') # get each line text on specific page's  
    checkpass = text[line].find(key) # get the result(pass or fail) poistion in result line  
    ''''''''''''''''''''''''''''''''''''''''''
    '''                end                 '''
    ''''''''''''''''''''''''''''''''''''''''''

    if checkpass > -1:
        return "Pass"
    else :
        return "Fail"

def ExcelFormat(wb):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    if "Sheet" in wb.sheetnames:
        wb["Sheet"].title = "USB2.0"

    ws = wb["USB2.0"]
    ws.column_dimensions['A'].width = 2.3
    ws.column_dimensions['B'].width = 3.6
    ws.column_dimensions['C'].width = 29.67
    ws.column_dimensions['D'].width = 17.33
    ws.column_dimensions['E'].width = 25.22
    ws.column_dimensions['F'].width = 29.65
    ws.column_dimensions['G'].width = 14.22
    ws.column_dimensions['H'].width = 28.67
    ws.column_dimensions['I'].width = 14.33
    ws.column_dimensions['J'].width = 11.89
    ws.column_dimensions['K'].width = 11.89
    ws.column_dimensions['L'].width = 11.89
    ws.column_dimensions['M'].width = 11.89
    ws.column_dimensions['N'].width = 11.89
    ws.column_dimensions['O'].width = 11.89
    ws.column_dimensions['P'].width = 11.89
    ws.column_dimensions['Q'].width = 11.89
    ws.column_dimensions['R'].width = 11.89

    #  set Unit Test 
    ws['A1'] = "Unit Test: USB Interface"
    ws['A1'].font = Font(name='Tahoma', size=16, bold=True, color="FFFFFF")
    ws.merge_cells('A1:R1')
    ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


    ws['A2'] = f'=SUM(O67:O999)/2' 
    ws['A3'] = f'=SUM(P67:P1000)/2' 
    ws['A4'] = f'=SUM(Q67:Q1000)/2' 
    ws['A5'] = f'=SUM(R67:R1002)/2' 
    ws['A2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells('A2:C2')
    ws.merge_cells('A3:C3')
    ws.merge_cells('A4:C4')
    ws.merge_cells('A5:C5')
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A5'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")


    ws['D2'] = "Tests Planned"
    ws['D3'] = "Tests Blocked"
    ws['D4'] = "Tests Passed"
    ws['D5'] = "Tests Failed"
    ws['D2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['D3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['D4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['D5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")    
    ws.merge_cells('D2:G2')
    ws.merge_cells('D3:G3')
    ws.merge_cells('D4:G4')
    ws.merge_cells('D5:G5')
    ws['D2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['D3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['D4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['D5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ###### UUT Information PART ########
    ws['A7'] = "UUT Information"
    ws['A7'].font = Font(name='Tahoma', size=12, bold=True, color="000000")
    ws['A7'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A8'] = "Project Name"
    ws['A8'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A9'] = "Project Revision"
    ws['A9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A10'] = "Part Number"
    ws['A10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A11'] = "Series Number"
    ws['A11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A12'] = "BIOS Version"
    ws['A12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A13'] = "BMC Version"
    ws['A13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A14'] = "CPU"
    ws['A14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A15'] = "PCH"
    ws['A15'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A16'] = "DIMM"
    ws['A16'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A17'] = "USB3.0 Cable (For JUSB3_INT1)"
    ws['A17'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A18'] = "USB2.0 Cable (For JUSB4 & JUSB5)"
    ws['A18'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A19'] = "Power supply"
    ws['A19'].font = Font(name='Tahoma', size=10, color="000000")    

    # let user input the UUT information 
    ws['D8'] = input("Project Name: ")
    ws['D8'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D9'] = input("Project Revision: ")
    ws['D9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D10'] = input("Part Number: ")
    ws['D10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D11'] = input("Series Number: ")
    ws['D11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D12'] = input("BIOS Version: ")
    ws['D12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D13'] = input("BMC Version: ")
    ws['D13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D14'] = input("CPU: ")
    ws['D14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D15'] = input("PCH: ")
    ws['D15'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D16'] = input("DIMM: ")
    ws['D16'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D17'] = input("USB3.0 Cable (For JUSB3_INT1): ")
    ws['D17'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D18'] = input("USB2.0 Cable (For JUSB4 & JUSB5): ")
    ws['D18'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D19'] = input("Power supply: ")
    ws['D19'].font = Font(name='Tahoma', size=10, color="000000")    
    # end of UUT input

    ws[f'A{7}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'A{8}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{9}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{10}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{11}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{12}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{13}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{14}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{15}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{16}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{17}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{18}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{19}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)    
    ws[f'D{7}'].border = Border(top=medium, left=thin, right=medium, bottom=thin)
    ws[f'D{8}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{9}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{10}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{11}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{12}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{13}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{14}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{15}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{16}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{17}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{18}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{19}'].border = Border(top=thin, left=thin, right=medium, bottom=medium) 
    
    ws.merge_cells('A7:G7')
    ws.merge_cells('A8:C8')
    ws.merge_cells('A9:C9')
    ws.merge_cells('A10:C10')
    ws.merge_cells('A11:C11')
    ws.merge_cells('A12:C12')
    ws.merge_cells('A13:C13')
    ws.merge_cells('A14:C14')
    ws.merge_cells('A15:C15')
    ws.merge_cells('A16:C16')
    ws.merge_cells('A17:C17')
    ws.merge_cells('A18:C18')
    ws.merge_cells('A19:C19')
    ws.merge_cells('D8:G8')
    ws.merge_cells('D9:G9')
    ws.merge_cells('D10:G10')
    ws.merge_cells('D11:G11')
    ws.merge_cells('D12:G12')
    ws.merge_cells('D13:G13')
    ws.merge_cells('D14:G14')
    ws.merge_cells('D15:G15')
    ws.merge_cells('D16:G16')
    ws.merge_cells('D17:G17')
    ws.merge_cells('D18:G18')
    ws.merge_cells('D19:G19')

    ##### TEST EQUIPMENT PART ###############
    ws['A22'] = "Test Equipment:"
    ws['A22'].font = Font(name='Tahoma', size=10, bold=True)

    ws['A23'] = "1."
    ws['A23'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A24'] = "2."
    ws['A24'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A25'] = "3."
    ws['A25'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A26'] = "4."
    ws['A26'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A27'] = "5."
    ws['A27'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A28'] = "6."
    ws['A28'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Equipment change '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B23'] = "Tektronix DSA72004C Digital Storage Oscilloscope 20GHz, 100GS/s"
    ws['B23'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B24'] = "Tektronix P6247 Differential Probe x 1"
    ws['B24'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B25'] = "Tektronix P6245 Active Probe x 2"
    ws['B25'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B26'] = "TDSUSBF USB2.0 Compliance Test Fixture With Power Adapter"
    ws['B26'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B27'] = "Digit Multimeter or above"
    ws['B27'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B28'] = "Fixture Solution USB 2.0, 3.X, BC1.2 Vbus Frop Droop Load Board"
    ws['B28'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ###### Test Auxiliary Application PART ##############
    ws['A31'] = "TEST AUXILIARY APPLICATION"
    ws['A31'].font = Font(name='Tahoma', size=10, bold=True)

    ws['A32'] = "1."
    ws['A32'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A33'] = "2."
    ws['A33'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Auxiliary Application change '''
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ws['B32'] = "XHCI_USBHSET_Releasex64_1.3.6.8"
    ws['B32'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B33'] = "TDSUSB2.0 Test Package Application (Ver 3.9.3 Build 5)"
    ws['B33'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ##### TEST PROCEDURE PART ###########
    ws['A36'] = "Test Procedure"
    ws['A36'].font = Font(name='Tahoma', size=10, bold=True)

    ws['A37'] = "1."
    ws['A37'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B37'] = "Droop :"
    ws['B37'].font = Font(name='Tahoma', size=10, color="000000")   
    ws['B38'] = "1.1"
    ws['B38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B39'] = "1.2"
    ws['B39'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B40'] = "1.3"
    ws['B40'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B41'] = "1.4"
    ws['B41'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B42'] = "1.5"
    ws['B42'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B43'] = "1.6"
    ws['B43'].font = Font(name='Tahoma', size=10, color="000000")

    ws['A44'] = "2."
    ws['A44'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B44'] = "Drop :"
    ws['B44'].font = Font(name='Tahoma', size=10, color="000000")   
    ws['B45'] = "2.1"
    ws['B45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B46'] = "2.2"
    ws['B46'].font = Font(name='Tahoma', size=10, color="000000")

    ws['A47'] = "3."
    ws['A47'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B47'] = "High Speed :"
    ws['B47'].font = Font(name='Tahoma', size=10, color="000000")   
    ws['B48'] = "3.1"
    ws['B48'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B49'] = "3.2"
    ws['B49'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B50'] = "3.3"
    ws['B50'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B51'] = "3.4"
    ws['B51'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B52'] = "3.5"
    ws['B52'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B53'] = "3.6"
    ws['B53'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if Test Procedure             '''
    '''''''''''''''''''''''''''''''''''''''''''''''''' 
    ws['C38'] = "Connect the host and the fixture with power adapter."
    ws['C38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C39'] = "Use the TDSUSBF USB2.0 compliance test fixture, connect two channels of Scope to the fixture."
    ws['C39'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C40'] = "Use switches to select 500mA load for each port except the nearest test port. "
    ws['C40'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C41'] = "Select Droop test from TDSUSB2.0 Test Package of application menu on Scope."
    ws['C41'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C42'] = "Observe droop voltage."
    ws['C42'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C43'] = "Repeat procedure for all ports."
    ws['C43'].font = Font(name='Tahoma', size=10, color="000000")

    ws['C45'] = "Use Multimeter to measure voltage for each port and record it. "
    ws['C45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C46'] = "Add 500mA load for each port, then to measure voltage and record it."
    ws['C46'].font = Font(name='Tahoma', size=10, color="000000")

    ws['C48'] = "Connect the host and the fixture with power adapter."
    ws['C48'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C49'] = "Use one differential probe to connect the Scope and the fixture."
    ws['C49'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C50'] = "Run ElectricalTestTool.EXE on host PC"
    ws['C50'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C51'] = "Run TDSUSB2 Application on Scope, select High Speed tab and configure the signal quality checks."
    ws['C51'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C52'] = "Run and output test reports."
    ws['C52'].font = Font(name='Tahoma', size=10, color="000000")
    ws['C53'] = "Repeat procedure for all ports."
    ws['C53'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ###### REF DOCUMENT PART #############
    ws['A56'] = "Ref Document"
    ws['A56'].font = Font(name='Tahoma', size=10, bold=True)

    ws['A57'] = "1."
    ws['A57'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A58'] = "2."
    ws['A58'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A59'] = "3."
    ws['A59'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if REF DOCUMENT change  '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B57'] = "USB Spec,  Rev 2.0, April 27, 2000"
    ws['B57'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B58'] = "USB 2.0 Pre-Compliance Physical Layer Testing"
    ws['B58'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B59'] = "USB-IF Full and Low Speed Electrical and Interoperability Compliance Test Procedure Revision 1.3, February 2004"
    ws['B59'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''


def Create_New_Excel(root):
    newwb = opxl.Workbook() # create a new excel with one default sheet, sheet name is "sheet"
    newwb.save(f'{root}\\Project_Signal-Integrity(USB2)_EVT_Test_Report_Rev.A0.xlsx') # name the excel "Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx"

def Get_Path():
    root = input("enter file path: ex.D:\SW1tool\QT_SI_Report_Auto\BP-PEX003-BD_Rev.A01\n") # user input the root file path
    # root = "D:\\SW1tool\\QT_SI_Report_Auto\\BP-PEX003-BD_Rev.A01_20240723"
    path = root + "\\Result_SI\\USB2.0_Tx"
    return path, root

def DroopHeader(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "Droop"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "Droop Voltage\n(mV)"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "Droop Voltage\nMax. Spec.\n(mV)"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "Pass/Fail"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:G{row+4}')
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+3}'] = "Notes"
    ws[f'H{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+3}:N{row+4}')
    ws[f'H{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def DropHeader(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'I{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)    
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'A{row}'] = "Drop"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "Vbus Without Loading\n(V)"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "USB\nVbus Without Loading Spec.  (V)"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "Vbus With\n500 mA\nLoading\n(V)"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:G{row+4}')
    ws[f'G{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+3}'] = "USB\nVbus With Loading Spec.  (V)"
    ws[f'H{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+3}:H{row+4}')
    ws[f'H{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'I{row+3}'] = "Notes"
    ws[f'I{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'I{row+3}:N{row+4}')
    ws[f'I{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def HighSpeedHeader(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "High Speed"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "Test Resuts"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "Pass/Fail"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "Notes"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:N{row+4}')
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def HighSpeed3334Header(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "High Speed Chrip EL_33,34 Timing"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "Test Resuts"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "Pass/Fail"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "Notes"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:N{row+4}')
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def HighSpeed35Header(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "High Speed Chrip EL_35 Timing"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "Test Resuts"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "Pass/Fail"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "Notes"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:N{row+4}')
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def USBOCPHeader(row, ws):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'I{row + 3}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)    
    ws[f'O{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'O{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'Q{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'R{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "USB OCP Test"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "USB Port"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:R{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Name"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:C{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "Probe Location"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'E{row+3}'] = "OCP SPEC\n(A)"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "OCP Load current\n(A)"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:F{row+4}')
    ws[f'F{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+3}'] = "OCP reaction\ntime"
    ws[f'G{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+3}:G{row+4}')
    ws[f'G{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'G{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+3}'] = "Pass/Fail"
    ws[f'H{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+3}:H{row+4}')
    ws[f'H{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'H{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'I{row+3}'] = "Notes"
    ws[f'I{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'I{row+3}:N{row+4}')
    ws[f'I{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'O{row + 3}'].value = "Number of Tests"
    ws[f'O{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'O{row+3}:R{row+3}')
    ws[f'O{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'O{row + 4}'].value = "Planned"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Blocked"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'Q{row + 4}'].value = "Passed"
    ws[f'Q{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'Q{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'Q{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'R{row + 4}'].value = "Failed"
    ws[f'R{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'R{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'R{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 5

def SignalName(ws, name, row, called):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    if called != "Drop":
        ws.row_dimensions[row].height = 54
    ws[f'A{row}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{row}'] = name
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

def ProbeLoc(ws, row):
    thin = Side(border_style="thin")
    ws[f'D{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'D{row}'] = "Tektronix Fixture"
    ws[f'D{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')

def ProbeLocUSB(ws, row):
    thin = Side(border_style="thin")
    ws[f'D{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'D{row}'] = "+5V_USB"
    ws[f'D{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')

def OCPSpec(ws, row):
    thin = Side(border_style="thin")
    ws[f'E{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'E{row}'] = "X.XX"
    ws[f'E{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'E{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')

def OCPReactTime(ws, row):
    thin = Side(border_style="thin")
    ws[f'G{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'G{row}'] = "Instantly"
    ws[f'G{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')    

def TestResult(row, filepath, ws, dict):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'E{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if filepath != None:
        dict[f'E{row}'] = filepath
    else:
        ws[f'E{row}'].value = "No Result file"
        ws[f'E{row}'].font = Font(name='Tahoma', size=10, color="FF0000")

    ws[f'E{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

def PassFail(column, row, result, ws):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    if result == "Fail":
        ws[f'{column}{row}'].value = result
        ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="FF0000")
        ws[f'{column}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    elif result == "Pass":
        ws[f'{column}{row}'].value = result
        ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")
        ws[f'{column}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True) 
    else:
        ws[f'{column}{row}'].value = "no result"
        ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="FF0000")
        ws[f'{column}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'{column}{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

def PassFailUSB(ws, row):
    thin = Side(border_style="thin")
    ws[f'H{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'H{row}'] = "Pass/Fail"
    ws[f'H{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')    

def TestData(row, data, ws):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'O{row}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'P{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'Q{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'R{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)

    ws[f'O{row}'].value = data[0]
    ws[f'O{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'O{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
  
    ws[f'P{row}'].value = data[1]
    ws[f'P{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'P{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
  
    ws[f'Q{row}'].value = data[2]
    ws[f'Q{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'Q{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    
    ws[f'R{row}'].value = data[3]
    ws[f'R{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'R{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    return row + 1

def DroopVoltageMax(ws, row):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'F{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'F{row}'] = "330.00"
    ws[f'F{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'F{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

def Notes(ws, row, column):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'{column}{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws.merge_cells(f'{column}{row}:N{row }')
    
def NumOfTests(row1, row2, ws):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    for i in col:
        ws[f'{i}{row2}'].border = Border(top=medium)


    ws[f'O{row2}'].border = Border(top=medium, left=medium, right=thin, bottom=medium)
    ws[f'P{row2}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'Q{row2}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'R{row2}'].border = Border(top=medium, left=thin, right=medium, bottom=medium)

    ws[f'N{row2}'] = 'Totals'
    ws[f'N{row2}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'N{row2}'].alignment = Alignment(horizontal="right", vertical="center", wrapText=True)
    
    ws[f'O{row2}'].value = f'=SUM(O{row1}:O{row2-1})'
    ws[f'O{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'O{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'P{row2}'].value = f'=SUM(P{row1}:P{row2-1})'
    ws[f'P{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'P{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'Q{row2}'].value = f'=SUM(Q{row1}:Q{row2-1})'
    ws[f'Q{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'Q{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'R{row2}'].value = f'=SUM(R{row1}:R{row2-1})'
    ws[f'R{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'R{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    return row2 + 3

def Dummy(ws, column, row):
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'{column}{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

def VbusWoLoad(ws, column, row):
    thin = Side(border_style="thin")
    ws[f'{column}{row}'] = "4.75 ~ 5.25"
    ws[f'{column}{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'{column}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'{column}{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

def Droop(file):
    drooplist = [] # pass or fial
    droopfile = [] # file path
    droopdata = [0, 0, 0, 0] # planned, blocked, passed, failed
    
    try:
        droopdata[0] += 1
        droop = GetPDFPASS(file)
    except: # block!!!
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", file)
        drooplist.append(file + " not found")
        droopfile.append(None)
        droopdata[1] += 1
    else:
        drooplist.append(droop)
        droopfile.append(file)
        if droop == "Pass":
            droopdata[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", file)
            droopdata[3] += 1

    return droopfile, drooplist, droopdata

def HighSpeed(file):
    highslist = [] # pass or fial
    highsfile = [] # file path
    highsdata = [0, 0, 0, 0] # planned, blocked, passed, failed

    try:
        highsdata[0] += 1
        highs = GetPDFPASS(file)
    except: # block!!!
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", file)
        highslist.append( file + " not found")
        highsfile.append(None)
        highsdata[1] += 1
    else:
        highslist.append(highs)
        highsfile.append(file)
        if highs == "Pass":
            highsdata[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", file)
            highsdata[3] += 1

    return highsfile, highslist, highsdata

def HighSpeed334(file):
    highs334list = [] # pass or fial
    highs334file = [] # file path
    highs334data = [0, 0, 0, 0] # planned, blocked, passed, failed

    try:
        highs334data[0] += 1
        highs334 = GetPDFPASS(file)
    except: # block!!!
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", file)
        highs334list.append(file + " not found")
        highs334file.append(None)
        highs334data[1] += 1
    else:
        highs334list.append(highs334)
        highs334file.append(file)
        if highs334 == "Pass":
            highs334data[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", file)
            highs334data[3] += 1
                        
    return highs334file, highs334list, highs334data

def HighSpeed35(file):
    highs35list = [] # pass or fial
    highs35file = [] # file path
    highs35data = [0, 0, 0, 0] # planned, blocked, passed, failed

    try:
        highs35data[0] += 1
        highs35 = GetPDFPASS(file)
    except: # block!!!
        print("!!!!!!!!!!!!!")
        print("!!! Block !!!")
        print("!!!!!!!!!!!!!")
        print("miss file: ", file)
        highs35list.append(file + " not found")
        highs35file.append(None)
        highs35data[1] += 1
    else:
        highs35list.append(highs35)
        highs35file.append(file)
        if highs35 == "Pass":
            highs35data[2] += 1
        else:
            print("!!!!!!!!!!!!!!")
            print("!!!  Fail  !!!")
            print("!!!!!!!!!!!!!!")
            print("Fail test result: ", file)
            highs35data[3] += 1

    return highs35file, highs35list, highs35data

def DroopResult(row, filepath, fileresult, data, ws, dicts):
    for i, file in enumerate(filepath):
        filename = os.path.basename(file)
        filename = filename[:filename.rfind('.')]
        SignalName(ws, filename, row, "Droop")
        ProbeLoc(ws, row)
        TestResult(row, filepath[i], ws, dicts)
        DroopVoltageMax(ws, row)
        PassFail('G', row, fileresult[i], ws)
        Notes(ws, row, 'H')
        row = TestData(row, data, ws)

    return row

def DropResult(row, filepath, ws):
    for file in filepath:
        if file.endswith(".pdf"): # only open the pdf:
            filename = os.path.basename(file)
            filename = filename[:filename.rfind('.')]
            SignalName(ws, filename, row, "Drop")
            ProbeLoc(ws, row)
            Dummy(ws, 'E', row)
            VbusWoLoad(ws, 'F', row)
            Dummy(ws, 'G', row)
            VbusWoLoad(ws, 'H', row)
            Notes(ws, row, 'I')
            data = [2, 0, 0, 0]
            row = TestData(row, data, ws)

    return row

def HighSpeedResult(row, filepath, fileresult, data, ws, dicts):
    for i, file in enumerate(filepath):
        filename = os.path.basename(file)
        filename = filename[:filename.rfind('.')]
        SignalName(ws, filename, row, "HS")
        ProbeLoc(ws, row)
        TestResult(row, filepath[i], ws, dicts)
        PassFail('F', row, fileresult[i], ws)
        Notes(ws, row, 'G')
        row = TestData(row, data, ws)
        
    return row

def HighSpeed334Result(row, filepath, fileresult, data, ws, dicts):
    for i, file in enumerate(filepath):
        filename = os.path.basename(file)
        filename = filename[:filename.rfind('.')]
        SignalName(ws, filename, row, "HS334")
        ProbeLoc(ws, row)
        TestResult(row, filepath[i], ws, dicts)
        PassFail('F', row, fileresult[i], ws)
        Notes(ws, row, 'G')
        row = TestData(row, data, ws)
        
    return row

def HighSpeed35Result(row, filepath, fileresult, data, ws, dicts):
    for i, file in enumerate(filepath):
        filename = os.path.basename(file)
        filename = filename[:filename.rfind('.')]
        SignalName(ws, filename, row, "HS334")
        ProbeLoc(ws, row)
        TestResult(row, filepath[i], ws, dicts)
        PassFail('F', row, fileresult[i], ws)
        Notes(ws, row, 'G')
        row = TestData(row, data, ws)
        
    return row

def USBOCPResult(row, filepath, ws):
    for  file in filepath:
        if file.endswith(".pdf"): # only open the pdf:
            filename = os.path.basename(file)
            filename = filename[:filename.rfind('.')]
            SignalName(ws, filename, row, "USBOCP")
            ProbeLocUSB(ws, row)
            OCPSpec(ws, row)
            Dummy(ws, 'F', row)
            OCPReactTime(ws, row)
            PassFailUSB(ws, row)
            Notes(ws, row, 'I')
            data = [1, 0, 0, 0]
            row = TestData(row, data, ws)

    return row

def Insertfile(dicts, root):
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    xl.Visible = False
    xl.ScreenUpdating = False
    xl.DisplayAlerts = False
    wb = xl.Workbooks.Open(f'{root}\\Project_Signal-Integrity(USB2)_EVT_Test_Report_Rev.A0.xlsx')
    ws = wb.Worksheets("USB2.0")

    for key in dicts:
        print("insert:", dicts[key])
        dest_cell = ws.Range(key)
        filename = dicts[key].split('\\')
        obj = ws.OLEObjects()
        obj.Add(Filename=dicts[key], Link=False, Left=dest_cell.Left, Top=dest_cell.Top, DisplayAsIcon=True, IconIndex=0, IconLabel=filename[-1], IconFileName="")
        obj.ShapeRange.LockAspectRatio = False
        obj.Height=53
        obj.Width=88
        print("===============================================================================================================")
    
    print("saving...")
    wb.Save()
    xl.Application.Quit()

def main():
    path, root = Get_Path() # root = path of project, path = path of project result
    os.chdir(root)
    Create_New_Excel(root)
    dirs = os.listdir(path)
    wb = opxl.load_workbook(f'{root}\\Project_Signal-Integrity(USB2)_EVT_Test_Report_Rev.A0.xlsx')
    ws = wb.active
    ExcelFormat(wb)
    dicts = {}
    tempdroopfile = []
    row = 62
    for dir in dirs:
        print("----------------------", dir, "--------------------------------")
        if dir == "Droop":
            row = DroopHeader(row, ws)
            row1 = row
            for _, _, files in os.walk(f'{path}\\{dir}'):
                tempdroopfile = files
                for file in files:
                    if file.endswith(".pdf"): # only open the pdf:
                        droopfile, drooplist, droopdata = Droop(f'{path}\\{dir}\\{file}')
                        row = DroopResult(row, droopfile, drooplist, droopdata, ws, dicts)

            row = NumOfTests(row1, row, ws)

            row = DropHeader(row, ws)
            row1 = row
            row = DropResult(row, tempdroopfile, ws)
            row = NumOfTests(row1, row, ws)

        elif dir == "High_Speed":
            row = HighSpeedHeader(row, ws)
            row1 = row
            for _, _, files in os.walk(f'{path}\\{dir}'):
                for file in files:
                    if file.endswith(".pdf"): # only open the pdf:
                        highsfile, highslist, highsdata = HighSpeed(f'{path}\\{dir}\\{file}')
                        row = HighSpeedResult(row, highsfile, highslist, highsdata, ws, dicts)

            row = NumOfTests(row1, row, ws)

        elif dir == "High_Speed_EL33.34":
            row = HighSpeed3334Header(row, ws)
            row1 = row
            for _, _, files in os.walk(f'{path}\\{dir}'):
                for file in files:
                    if file.endswith(".pdf"): # only open the pdf:
                        highs334file, highs334list, highs334data = HighSpeed334(f'{path}\\{dir}\\{file}')
                        row = HighSpeed334Result(row, highs334file, highs334list, highs334data, ws, dicts)

            row = NumOfTests(row1, row, ws)

        elif dir == "High_Speed_EL35":
            row = HighSpeed35Header(row, ws)
            row1 = row            
            for _, _, files in os.walk(f'{path}\\{dir}'):
                for file in files:
                    if file.endswith(".pdf"): # only open the pdf:
                        highs35file, highs35list, highs35data = HighSpeed35(f'{path}\\{dir}\\{file}')
                        row = HighSpeed35Result(row, highs35file, highs35list, highs35data, ws, dicts)

            row = NumOfTests(row1, row, ws)  

            row = USBOCPHeader(row, ws)
            row1 = row
            row = USBOCPResult(row, tempdroopfile, ws)
            row = NumOfTests(row1, row, ws)
    
    wb.save(f'{root}\\Project_Signal-Integrity(USB2)_EVT_Test_Report_Rev.A0.xlsx')

    Insertfile(dicts, root)
    print("============ DONE ============")


if __name__ == '__main__':
    main()
