import cv2
import os
import re
import openpyxl as opxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from paddleocr import PaddleOCR

'''
### 第一次執行完，請手動修改det、rec、cls路徑、取消註解此區，並將原code註解
### need to download modle manually, https://github.com/PaddlePaddle/PaddleOCR/blob/release/2.7/doc/doc_ch/models_list.md#1.1
### find ch_PP-OCRv4_det, ch_PP-OCRv3_rec, ch_ppocr_mobile_v2.0_cls then download them
### unzip them and place them in .paddleocr as below
det = "C:\\Users\\sam.lu\\.paddleocr\\whl\\det\\ch\\ch_PP-OCRv4_det_infer" # change detection model path of your own 
rec = "C:\\Users\\sam.lu\\.paddleocr\\whl\\rec\\ch\\ch_PP-OCRv3_rec_infer" # change recognition model path of your own 
cls = "C:\\Users\\sam.lu\\.paddleocr\\whl\\cls\\ch_ppocr_mobile_v2.0_cls_infer" # change classification model path of your own 

# init OCR
ocr = PaddleOCR(
    lang="ch",
    det_model_dir=det,
    rec_model_dir=rec,
    cls_model_dir=cls,
)
'''

ocr = PaddleOCR(lang='ch') # 我是原code，第一次執行完把我註解

want = ["Maximum", "Minimum", "Rise Time", "Fall Time", "Frequency", "High Time", "Low Time", "Hold Time", "Setup Time"]

def OCR(imagepath, dict1, dict2):
    print(f'{imagepath} is OCRing...')
    img = cv2.imread(imagepath)
    # OCR results
    result = ocr.ocr(img, cls=False)
    # Dictionary to store the desired results
    # Print results and filter based on desired labels and position
    for i, res in enumerate(result[0]):
        print(res)
        if "vihvil" in imagepath and res[1][0] in want:
            if result[0][i+1][1][0] == "Ch 1" or result[0][i+1][1][0] == "ch 1" or\
               result[0][i+1][1][0] == "Ch 1, Ch 2" or result[0][i+1][1][0] == "ch 1, ch 2" or\
               result[0][i+1][1][0] == "Ch 1, ch 2" or result[0][i+1][1][0] == "ch 1, Ch 2":
                dict1[res[1][0]] = result[0][i+2][1][0]
            elif result[0][i+1][1][0] == "Ch 2" or result[0][i+1][1][0] == "ch 2" or\
                 result[0][i+1][1][0] == "Ch 2, Ch 1" or result[0][i+1][1][0] == "ch 2, ch 1" or\
                 result[0][i+1][1][0] == "Ch 2, ch 1" or result[0][i+1][1][0] == "ch 2, Ch 1":
                dict2[res[1][0]] = result[0][i+2][1][0]
        elif "setuptime" in imagepath and res[1][0] == "Setup Time":
            if result[0][i+1][1][0] == "Ch 2, Ch 1" or result[0][i+1][1][0] == "ch 2, ch 1" or\
               result[0][i+1][1][0] == "Ch 2, ch 1" or result[0][i+1][1][0] == "ch 2, Ch 1":
                dict2[res[1][0]] = result[0][i+2][1][0]

def ExcelFormat(wb):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    if "Sheet" in wb.sheetnames:
        wb["Sheet"].title = "I2C_P"

    ws = wb["I2C_P"]
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 19
    ws.column_dimensions['D'].width = 23
    ws.column_dimensions['E'].width = 23
    ws.column_dimensions['F'].width = 23
    ws.column_dimensions['G'].width = 23
    ws.column_dimensions['H'].width = 23
    ws.column_dimensions['I'].width = 23
    ws.column_dimensions['J'].width = 23
    ws.column_dimensions['K'].width = 23
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 12


    # 設定標題
    ws['A1'] = "Unit Test: I2C"
    ws['A1'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # 合併單元格
    ws.merge_cells('A1:P1')

    ws['A2'] = f'=SUM(L28:L809)/2' 
    ws['A3'] = f'=SUM(M28:M809)/2' 
    ws['A4'] = f'=SUM(N28:N809)/2' 
    ws['A5'] = f'=SUM(O28:O809)/2' 
    ws['A6'] = f'=SUM(P61:P810)/2'


    ws.merge_cells('A2:B2')
    ws.merge_cells('A3:B3')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A5:B5')
    ws.merge_cells('A6:B6')

    ws['C2'] = "Tests Planned"
    ws['C3'] = "Tests Blocked"
    ws['C4'] = "Tests Passed"
    ws['C5'] = "Tests N/A"
    ws['C6'] = "Tests Failed"

    ws.merge_cells('C2:G2')
    ws.merge_cells('C3:G3')
    ws.merge_cells('C4:G4')
    ws.merge_cells('C5:G5')
    ws.merge_cells('C6:G6')
    
    ws['A2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['A6'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C2'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C3'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C4'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C5'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws['C6'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")

    ws['C2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['C6'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")


    ws['A2'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A3'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A4'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A5'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    ws['A6'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A5'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws['A6'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


    ###### UUT PART ########
    ws['A8'] = "UUT Information"
    ws['A8'].font = Font(name='Tahoma', size=12, bold=True, color="000000")
    ws['A8'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A9'] = "Project Name"
    ws['A9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A10'] = "Project Revision"
    ws['A10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A11'] = "Part Number"
    ws['A11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A12'] = "Series Number"
    ws['A12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A13'] = "PEX89114 BIOS Version"
    ws['A13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A14'] = "COM Express"
    ws['A14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A15'] = "DIMM"
    ws['A15'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A16'] = "Power supply"
    ws['A16'].font = Font(name='Tahoma', size=10, color="000000")    

    # let user input the UUT information 
    ws['D9'] = input("enter Project Name: ")
    ws['D9'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D10'] = input("enter Project Revision: ")
    ws['D10'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D11'] = input("enter Part Number: ")
    ws['D11'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D12'] = input("enter Series Number: ")
    ws['D12'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D13'] = input("enter PEX89144 BIOS Version: ")
    ws['D13'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D14'] = input("enter COM Express: ")
    ws['D14'].font = Font(name='Tahoma', size=10, color="000000")
    ws['D15'] = input("enter DIMM: ")
    ws['D15'].font = Font(name='Tahoma', size=10, color="000000")   
    ws['D16'] = input("Power supply: ")
    ws['D16'].font = Font(name='Tahoma', size=10, color="000000")    
    # end of UUT input

    ws[f'A{8}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)
    ws[f'A{9}'].border = Border(top=medium, left=medium, right=medium, bottom=thin)
    ws[f'A{10}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{11}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{12}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{13}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{14}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{15}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'A{16}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'D{9}'].border = Border(top=medium, left=thin, right=medium, bottom=thin)
    ws[f'D{10}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{11}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{12}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{13}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{14}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{15}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws[f'D{16}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws.merge_cells('A8:G8')
    ws.merge_cells('A9:C9')
    ws.merge_cells('A10:C10')
    ws.merge_cells('A11:C11')
    ws.merge_cells('A12:C12')
    ws.merge_cells('A13:C13')
    ws.merge_cells('A14:C14')
    ws.merge_cells('A15:C15')
    ws.merge_cells('A16:C16')

    ws.merge_cells('D9:G9')
    ws.merge_cells('D10:G10')
    ws.merge_cells('D11:G11')
    ws.merge_cells('D12:G12')
    ws.merge_cells('D13:G13')
    ws.merge_cells('D14:G14')
    ws.merge_cells('D15:G15')
    ws.merge_cells('D16:G16')

    ##### TEST EQUIPMENT PART ###############
    ws['A18'] = "Test Equipment:"
    ws['A18'].font = Font(name='Tahoma', size=10, bold=True)
    ws['A19'] = "1."
    ws['A19'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A20'] = "2."
    ws['A20'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Equipment change '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B19'] = "Tektronix DPO7254 Digital Storage Oscilloscope 2.5GHz, 40GS/s"
    ws['B19'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B20'] = "Tektronix P6245 Active Singel end Probe x 2"
    ws['B20'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''


    ###### Test Auxiliary Application PART ##############
    ws['A23'] = "Test Condition"
    ws['A23'].font = Font(name='Tahoma', size=10, bold=True)
    ws['A24'] = "1."
    ws['A24'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A25'] = "2."
    ws['A25'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A26'] = "3."
    ws['A26'].font = Font(name='Tahoma', size=10, color="000000")

    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ''' modify here if Test Auxiliary Application change '''
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    ws['B24'] = "Rise Time Limits: 0.3 *VDD to 0.7 *VDD"
    ws['B24'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B25'] = "Fall Time Limits: 0.7 *VDD to 0.3 *VDD"
    ws['B25'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B26'] = "Cursor ->0.3 *VDD ~ 0.7 *VDD"
    ws['B26'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ##### TEST PROCEDURE PART ###########
    ws['A29'] = "Test Procedure"
    ws['A30'].font = Font(name='Tahoma', size=10, bold=True)
    ws['A31'] = "1."
    ws['A31'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A32'] = "2."
    ws['A32'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A33'] = "3."
    ws['A33'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A34'] = "4."
    ws['A34'].font = Font(name='Tahoma', size=10, color="000000")


    ''''''''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if Test Procedure - TX change '''
    ''''''''''''''''''''''''''''''''''''''''''''''''''
    ws['B30'] = "Warm up Scope 20 mins at least, and calibrate Probes. "
    ws['B30'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B31'] = "Select the longest, shortest and middle traces for measurement. "
    ws['B31'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B32'] = "According to the signal type, probing the signal-end probe to the test point to use a minimum loop area methodology."
    ws['B32'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B33'] = "Capture result of setup time, hold time, Vih, Vil, overshoot and undershoot on the scope.  "
    ws['B33'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

    ###### REF DOCUMENT PART #############
    ws['A36'] = "Ref Document"
    ws['A36'].font = Font(name='Tahoma', size=10, bold=True)
    ws['A37'] = "1."
    ws['A37'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A38'] = "2."
    ws['A38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A39'] = "3."
    ws['A39'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A40'] = "4."
    ws['A40'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A41'] = "5."
    ws['A41'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A42'] = "6."
    ws['A42'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A43'] = "7."
    ws['A43'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A44'] = "8."
    ws['A44'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A45'] = "9."
    ws['A45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A46'] = "10."
    ws['A46'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A47'] = "11."
    ws['A47'].font = Font(name='Tahoma', size=10, color="000000")
    ws['A48'] = "12."
    ws['A48'].font = Font(name='Tahoma', size=10, color="000000")



    ''''''''''''''''''''''''''''''''''''''''''''
    '''  modify here if REF DOCUMENT change  '''
    ''''''''''''''''''''''''''''''''''''''''''''
    ws['B37'] = "Aspeed AST2500 Integrated Remote Management Processor A2 Datasheet Rev.1.6, Mat, 2017."
    ws['B37'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B38'] = "I2C: UM10204 I2C-bus specification and user manual, Rev.6, April 4, 2014."
    ws['B38'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B39'] = "Pmbus:CM1-00007000.pdf_PWR CTRL_MP2985BGLUT, Rev0.1, 6/6/2023"
    ws['B39'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B40'] = "Pmbus:CM1-00007003.pdf_PWR CTRL_MP2992BGMKT, Rev0.1, 5/26/2023"
    ws['B40'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B41'] = "Broadcom PEX89144 PCIe 5.0 Switch Data Sheet Rev.2.1."
    ws['B41'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B42'] = "NXP Semiconductors PCA9541A 2-to-1 I2C-bus master selector with interrupt logic and reset Rev.5, Apr, 2014."
    ws['B42'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B43'] = "NXP Semiconductors PCA9546A 4-channel I2C-bus switch with reset Rev.6, Apr, 2014"
    ws['B43'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B44'] = "NXP Semiconductors PCA9554; PCA9554A 8-bit I2C-bus and SMBus I/O port with interrupt Rev.10, Nov, 2017."
    ws['B44'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B45'] = "NXP Semiconductors PCA9555 16-bit I2C-bus and SMBus I/O port with interrupt Rev.10, Nov, 2017.."
    ws['B45'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B46'] = "Texas Instruments LM95241 Dual Remote Diode Temperature Sensor with SMBus Interface and TruTherm™ Technology (65nm/90nm), Mar, 2013."
    ws['B46'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B47'] = "Atmel AT24C128C I2C-Compatible (2-Wire) Serial EEPROM 128-Kbit (16,384 x 8)."
    ws['B47'].font = Font(name='Tahoma', size=10, color="000000")
    ws['B48'] = "MonolithicPower MP5023 16V, 50A, 1.1mΩ, Protection Device with Integrated MOSFET and PMBusTM Interface Rev.1.12, Jan, 2019."
    ws['B48'].font = Font(name='Tahoma', size=10, color="000000")
    ''''''''''''''''''''''''''''''''''''''''''''
    '''               end                    '''
    ''''''''''''''''''''''''''''''''''''''''''''

def Create_New_Excel(root):
    newwb = opxl.Workbook() # create a new excel with one default sheet, sheet name is "sheet"
    newwb.save(f'{root}\\Project_Signal-Integrity(I2C_P)_EVT_Test_Report_Rev.A0.xlsx') # name the excel "Project_Signal-Integrity_EVT_Test_Report_Rev.A0.xlsx"

def Get_Path():
    print()
    print()
    root = input("enter file path: ex.D:\\SW1tool\\QT_SI_Report_Auto\\BP-PEX003-BD_Rev.A01\\ISC\n") # user input the root file path
    # root = r"D:\SW1tool\QT_SI_Report_Auto\BP-PEX003-BD_Rev.A01_20240723\Result_SI\I2C\BMC\I2C2"
    path = root + "\\Result_SI\\I2C_JohnP"
    return root, path

def VDDVol(ws, row):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row + 2}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)
    ws[f'A{row + 3}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'D{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 3}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'L{row + 3}'].border = Border(top=thin, left=medium, right=medium, bottom=medium)

    ws[f'L{row + 4}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row + 4}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row + 4}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)

    ws[f'A{row}'] = "VDD Voltage Test"
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    ws[f'A{row+2}'] = "Voltage Quality"
    ws[f'A{row+2}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row+2}:P{row+2}')
    ws[f'A{row+2}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row + 3].height = 14.1
    ws.row_dimensions[row + 4].height = 60.2

    ws[f'A{row+3}'] = "Signal Pre Pin"
    ws[f'A{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+3}:B{row+4}')
    ws[f'A{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'C{row+3}'] = "Probe Location"
    ws[f'C{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'C{row+3}:C{row+4}')
    ws[f'C{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+3}'] = "VDD\n(mV)"
    ws[f'D{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+3}:D{row+4}')
    ws[f'D{row+3}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'E{row+3}'] = "VDD\nSpec.\n(V)"
    ws[f'E{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+3}:E{row+4}')
    ws[f'E{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+3}'] = "Notes"
    ws[f'F{row+3}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+3}:K{row+4}')
    ws[f'F{row+3}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'L{row + 3}'].value = "Number of Tests"
    ws[f'L{row + 3}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'L{row+3}:P{row+3}')
    ws[f'L{row + 3}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 3}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'L{row + 4}'].value = "Planned"
    ws[f'L{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'L{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 4}'].value = "Blocked"
    ws[f'M{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'M{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'M{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'N{row + 4}'].value = "Passed"
    ws[f'N{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'N{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'N{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'O{row + 4}'].value = "N/A"
    ws[f'O{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 4}'].value = "Failed"
    ws[f'P{row + 4}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 4}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 4}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'A{row + 5}'].border = Border(top=medium, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 5}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    ws[f'D{row + 5}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)   
    ws[f'E{row + 5}'].border = Border(top=medium, left=thin, right=thin, bottom=medium)

    ws.merge_cells(f'A{row+5}:B{row+5}')

    ws[f'E{row+5}'] = "2 ~ 3.6"
    ws[f'E{row+5}'].font = Font(name='Tahoma', size=10)
    ws[f'E{row+5}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'E{row+5}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    Notes(ws, row+5, 'F')
    TestData(row+5, [1, 0, 1, 0, 0], ws)
    NumOfTests(row+5, row+6, ws)

    return row + 9

def DirHeader(ws, row, name):
    ws[f'A{row}'] = f'{name}(0xXX) Signal Integrity - I2C'
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    return row + 2

def Header(ws, row, name):
    ws[f'A{row}'] = name
    ws[f'A{row}'].font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="0070c0", end_color="0070c0", fill_type="solid")

    return row + 1

def VolatageHeader(ws, row):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)    
    ws[f'D{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'L{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)    

    ws[f'L{row + 2}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'A{row}'] = "Voltage Quality"
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row+1].height = 14.1
    ws.row_dimensions[row+2].height = 60.2

    ws[f'A{row+1}'] = "Signal Pre Pin"
    ws[f'A{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+1}:B{row+2}')
    ws[f'A{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'C{row+1}'] = "Probe Location"
    ws[f'C{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'C{row+1}:C{row+2}')
    ws[f'C{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+1}'] = "VIH\n(V)"
    ws[f'D{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+1}:D{row+2}')
    ws[f'D{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'E{row+1}'] = "VIH\nSpec.\n(V)"
    ws[f'E{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+1}:E{row+2}')
    ws[f'E{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+1}'] = "VIL\n(mV)"
    ws[f'F{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+1}:F{row+2}')
    ws[f'F{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+1}'] = "VIL\nSpec.\n(V)"
    ws[f'G{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+1}:G{row+2}')
    ws[f'G{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+1}'] = "Notes"
    ws[f'H{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+1}:K{row+2}')
    ws[f'H{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'L{row + 1}'].value = "Number of Tests"
    ws[f'L{row + 1}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'L{row+1}:P{row+1}')
    ws[f'L{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'L{row + 2}'].value = "Planned"
    ws[f'L{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'L{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 2}'].value = "Blocked"
    ws[f'M{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'M{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'M{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'N{row + 2}'].value = "Passed"
    ws[f'N{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'N{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'N{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'O{row + 2}'].value = "N/A"
    ws[f'O{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 2}'].value = "Failed"
    ws[f'P{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 3

def RiseFallTimeHeader(ws, row):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)    
    ws[f'D{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'L{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)    

    ws[f'L{row + 2}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'A{row}'] = "Rise/Fall Time"
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row+1].height = 14.1
    ws.row_dimensions[row+2].height = 60.2

    ws[f'A{row+1}'] = "Signal Name"
    ws[f'A{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+1}:B{row+2}')
    ws[f'A{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'C{row+1}'] = "Probe Location"
    ws[f'C{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'C{row+1}:C{row+2}')
    ws[f'C{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+1}'] = "Rise Time\n(ns)"
    ws[f'D{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+1}:D{row+2}')
    ws[f'D{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'E{row+1}'] = "Rise Time\nMax. Spec.\n(ns)"
    ws[f'E{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+1}:E{row+2}')
    ws[f'E{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+1}'] = "Fall Time\n(ns)"
    ws[f'F{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+1}:F{row+2}')
    ws[f'F{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+1}'] = "Fall Time\nMax. Spec.\n(ns)"
    ws[f'G{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+1}:G{row+2}')
    ws[f'G{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+1}'] = "Notes"
    ws[f'H{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+1}:K{row+2}')
    ws[f'H{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'L{row + 1}'].value = "Number of Tests"
    ws[f'L{row + 1}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'L{row+1}:P{row+1}')
    ws[f'L{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'L{row + 2}'].value = "Planned"
    ws[f'L{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'L{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 2}'].value = "Blocked"
    ws[f'M{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'M{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'M{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'N{row + 2}'].value = "Passed"
    ws[f'N{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'N{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'N{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'O{row + 2}'].value = "N/A"
    ws[f'O{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 2}'].value = "Failed"
    ws[f'P{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 3

def CLKQualityHeader(ws, row):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)    
    ws[f'D{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'I{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'J{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'L{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)    

    ws[f'L{row + 2}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'A{row}'] = "CLK Quality"
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row+1].height = 14.1
    ws.row_dimensions[row+2].height = 60.2

    ws[f'A{row+1}'] = "Signal Name"
    ws[f'A{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+1}:B{row+2}')
    ws[f'A{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'C{row+1}'] = "Probe Location"
    ws[f'C{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'C{row+1}:C{row+2}')
    ws[f'C{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+1}'] = "Frequence\n(kHz)"
    ws[f'D{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+1}:D{row+2}')
    ws[f'D{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'E{row+1}'] = "Frequence\nSpec.\n(kHz)"
    ws[f'E{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+1}:E{row+2}')
    ws[f'E{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+1}'] = "THIGH\n(us)"
    ws[f'F{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+1}:F{row+2}')
    ws[f'F{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+1}'] = "THIGH\nMin. Spec.\n(us)"
    ws[f'G{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+1}:G{row+2}')
    ws[f'G{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+1}'] = "TLOW\n(us)"
    ws[f'H{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+1}:H{row+2}')
    ws[f'H{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'H{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'I{row+1}'] = "TLOW\nMin. Spec.\n(us)"
    ws[f'I{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'I{row+1}:I{row+2}')
    ws[f'I{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'J{row+1}'] = "Notes"
    ws[f'J{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'J{row+1}:K{row+2}')
    ws[f'J{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'L{row + 1}'].value = "Number of Tests"
    ws[f'L{row + 1}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'L{row+1}:P{row+1}')
    ws[f'L{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'L{row + 2}'].value = "Planned"
    ws[f'L{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'L{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 2}'].value = "Blocked"
    ws[f'M{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'M{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'M{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'N{row + 2}'].value = "Passed"
    ws[f'N{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'N{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'N{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'O{row + 2}'].value = "N/A"
    ws[f'O{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 2}'].value = "Failed"
    ws[f'P{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 3

def SetupHoldTimmingHeader(ws, row):
    thin = Side(border_style="thin")
    medium = Side(border_style="medium")

    ws[f'A{row}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)

    ws[f'A{row + 1}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'C{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)    
    ws[f'D{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'E{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'F{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'G{row + 1}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'H{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'I{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'J{row + 1}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)
    ws[f'L{row + 1}'].border = Border(top=medium, left=medium, right=medium, bottom=medium)       

    ws[f'L{row + 2}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row + 2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row + 2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'A{row}'] = "Setup/Hold Timing"
    ws[f'A{row}'].font = Font(name='Tahoma', size=10, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.row_dimensions[row+1].height = 14.1
    ws.row_dimensions[row+2].height = 60.2

    ws[f'A{row+1}'] = "Signal Name"
    ws[f'A{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'A{row+1}:B{row+2}')
    ws[f'A{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'C{row+1}'] = "Probe Location"
    ws[f'C{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'C{row+1}:C{row+2}')
    ws[f'C{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center")

    ws[f'D{row+1}'] = "tSU\nSetup Time\n(ns)"
    ws[f'D{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'D{row+1}:D{row+2}')
    ws[f'D{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'D{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'E{row+1}'] = "tSU\nSetup Time\nMin. Spec.\n(ns)"
    ws[f'E{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'E{row+1}:E{row+2}')
    ws[f'E{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'F{row+1}'] = "tHD\nHold Time\n(ns)"
    ws[f'F{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'F{row+1}:F{row+2}')
    ws[f'F{row+1}'].fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    ws[f'F{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'G{row+1}'] = "tHD\nHold Time\nMin. Spec.\n(us)"
    ws[f'G{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'G{row+1}:G{row+2}')
    ws[f'G{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'H{row+1}'] = "Notes"
    ws[f'H{row+1}'].font = Font(name='Tahoma', size=10, bold=True, color="000000")
    ws.merge_cells(f'H{row+1}:K{row+2}')
    ws[f'H{row+1}'].alignment = Alignment(vertical="bottom", horizontal="center", wrapText=True)

    ws[f'L{row + 1}'].value = "Number of Tests"
    ws[f'L{row + 1}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'L{row+1}:P{row+1}')
    ws[f'L{row + 1}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 1}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'L{row + 2}'].value = "Planned"
    ws[f'L{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'L{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'L{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'M{row + 2}'].value = "Blocked"
    ws[f'M{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'M{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'M{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'N{row + 2}'].value = "Passed"
    ws[f'N{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'N{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'N{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'O{row + 2}'].value = "N/A"
    ws[f'O{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'O{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'O{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'P{row + 2}'].value = "Failed"
    ws[f'P{row + 2}'].font = Font(name='Tahoma', size=10)
    ws[f'P{row + 2}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws[f'P{row + 2}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    return row + 3

def Notes(ws, row, column):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")
    ws[f'{column}{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'{column}{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)
    ws.merge_cells(f'{column}{row}:K{row }')

def TestData(row, data, ws):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'L{row}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'M{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'N{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'O{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'P{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)

    ws[f'L{row}'].value = data[0]
    ws[f'L{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'L{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
  
    ws[f'M{row}'].value = data[1]
    ws[f'M{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'M{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
  
    ws[f'N{row}'].value = data[2]
    ws[f'N{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'N{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    
    ws[f'O{row}'].value = data[3]
    ws[f'O{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'O{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    ws[f'P{row}'].value = data[4]
    ws[f'P{row}'].font = Font(name='Tahoma', size=10, color="000000")
    ws[f'P{row}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    return row + 1

def NumOfTests(row1, row2, ws):
    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for i in col:
        ws[f'{i}{row2}'].border = Border(top=medium)


    ws[f'L{row2}'].border = Border(top=thin, left=medium, right=thin, bottom=medium)
    ws[f'M{row2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'N{row2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'O{row2}'].border = Border(top=thin, left=thin, right=thin, bottom=medium)
    ws[f'P{row2}'].border = Border(top=thin, left=thin, right=medium, bottom=medium)


    ws[f'K{row2}'] = 'Totals'
    ws[f'K{row2}'].font = Font(name='Tahoma', size=10, bold =True)
    ws[f'K{row2}'].alignment = Alignment(horizontal="right", vertical="center", wrapText=True)
    
    ws[f'L{row2}'].value = f'=SUM(L{row1}:L{row2-1})'
    ws[f'L{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'L{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'M{row2}'].value = f'=SUM(M{row1}:M{row2-1})'
    ws[f'M{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'M{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'N{row2}'].value = f'=SUM(N{row1}:N{row2-1})'
    ws[f'N{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'N{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'O{row2}'].value = f'=SUM(O{row1}:O{row2-1})'
    ws[f'O{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'O{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    ws[f'P{row2}'].value = f'=SUM(P{row1}:P{row2-1})'
    ws[f'P{row2}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'P{row2}'].font = Font(name='Tahoma', size=10, bold =True)

    return row2 + 4

def Result_1(ws, row, name, Probe, result1, spec1, result2, spec2, column, data, misslist, needloc=False):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'C{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'D{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'E{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'F{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'G{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)

    locname = f'A{row}'
    ws[f'A{row}'] = name
    ws[f'A{row}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    locprobe = f'C{row}'
    ws[f'C{row}'] = Probe
    ws[f'C{row}'].font = Font(name='Tahoma', size=10)
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'D{row}'] = result1
    ws[f'D{row}'].font = Font(name='Tahoma', size=10)
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    if result1 == None:
        misslist.append(f'D{row}')

    ws[f'E{row}'] = spec1
    ws[f'E{row}'].font = Font(name='Tahoma', size=10)
    ws[f'E{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'E{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'F{row}'] = result2
    ws[f'F{row}'].font = Font(name='Tahoma', size=10)
    ws[f'F{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    if result2 == None:
        misslist.append(f'F{row}')  

    ws[f'G{row}'] = spec2
    ws[f'G{row}'].font = Font(name='Tahoma', size=10)
    ws[f'G{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'G{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    Notes(ws, row, column)
    row = TestData(row, data, ws)
    
    if needloc:
        return row, f'={locname}', f'={locprobe}', misslist
    
    return row, misslist

def Result_2(ws, row, name, Probe, result1, spec1, result2, spec2, result3, spec3, column, data, misslist):

    medium = Side(border_style="medium")
    thin = Side(border_style="thin")

    ws[f'A{row}'].border = Border(top=thin, left=medium, right=thin, bottom=thin)
    ws[f'C{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'D{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'E{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'F{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'G{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'H{row}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)    
    ws[f'I{row}'].border = Border(top=thin, left=thin, right=medium, bottom=thin)


    ws[f'A{row}'] = name
    ws[f'A{row}'].font = Font(name='Tahoma', size=10)
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'C{row}'] = Probe
    ws[f'C{row}'].font = Font(name='Tahoma', size=10)
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'D{row}'] = result1
    ws[f'D{row}'].font = Font(name='Tahoma', size=10)
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)
    
    if result1 == None:
        misslist.append(f'D{row}')

    ws[f'E{row}'] = spec1
    ws[f'E{row}'].font = Font(name='Tahoma', size=10)
    ws[f'E{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'E{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'F{row}'] = result2
    ws[f'F{row}'].font = Font(name='Tahoma', size=10)
    ws[f'F{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    if result2 == None:
        misslist.append(f'F{row}')

    ws[f'G{row}'] = spec2
    ws[f'G{row}'].font = Font(name='Tahoma', size=10)
    ws[f'G{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'G{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    ws[f'H{row}'] = result3
    ws[f'H{row}'].font = Font(name='Tahoma', size=10)
    ws[f'H{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    if result3 == None:
        misslist.append(f'H{row}')

    ws[f'I{row}'] = spec3
    ws[f'I{row}'].font = Font(name='Tahoma', size=10)
    ws[f'I{row}'].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws[f'I{row}'].alignment = Alignment(horizontal="center", vertical="bottom", wrapText=True)

    Notes(ws, row, column)
    row = TestData(row, data, ws)

    return row, misslist

def ListSubFolders(directory):
    subfolders = []
    for entry in os.scandir(directory):
        if entry.is_dir():
            subfolders.append(entry.path)
    return subfolders

def time_to_ns(time_value, unit):
    conversion_factors = {
        's': 1e9,
        'ms': 1e6,
        'μs': 1e3,
        'us': 1e3,
        'ns': 1e0  
    }
    return ('%.3f'%(time_value * conversion_factors[unit]))

def time_to_us(time_value, unit):
    conversion_factors = {
        's': 1e6,
        'ms': 1e3,
        'μs': 1e0,
        'us': 1e0,
        'ns': 1e-3
    }
    return ('%.3f'%(time_value * conversion_factors[unit]))

def voltage_to_mv(voltage_value, unit):
    conversion_factors = {
        'V': 1e3,
        'mV': 1e0,
        'μV': 1e-3,
        'uV': 1e-3,
        'nV': 1e-6
    }
    return ('%.3f'%(voltage_value * conversion_factors[unit]))

def voltage_to_v(voltage_value, unit):
    conversion_factors = {
        'V': 1e0,
        'mV': 1e-3,
        'μV': 1e-6,
        'uV': 1e-6,
        'nV': 1e-9
    }
    return ('%.3f'%(voltage_value * conversion_factors[unit]))

def frequency_to_khz(frequency_value, unit):
    conversion_factors = {
        'kHz': 1e0,
        'Hz': 1e-3,
        'mHz': 1e-6,
        'μHz': 1e-9,
        'uHz': 1e-9,
        'nV': 1e-12  
    }
    return ('%.3f'%(frequency_value * conversion_factors[unit]))

def GetValueUnit(value):
    ### after observe the images, only below units are legal, the others will be put into wrong text then return none 
    match = re.match(r"([-]?[0-9.]+) (nV|μV|uV|mV|V|nHz|μHz|uHz|mHz|Hz|kHz|ns|μs|us|ms|s|%)", value)
    if match:
        value = float(match.group(1))
        unit = match.group(2)
        return value, unit
    else:
        return None, None

def UnitTransform(dict):
    ### infor 1. this program use GetValueUnit to check if get the correct text, more information please see GetValueUnit
    ### infor 2. see GetValueUnit
    ### infor 3. after GetValueUnit, if get wrong text, the value will get None, then dict delete the key. In other words, this program did not detect it.
    v = ["Maximum"]
    mv = ["Minimum"]
    ns = ["Rise Time", "Fall Time", "Hold Time", "Setup Time"]
    us = ["High Time", "Low Time"]
    khz = ["Frequency"]

    for key in dict:
        value, unit = GetValueUnit(dict[key])
        if value == None:
            dict[key] = None
            continue
        
        if key in v:
            dict[key] = voltage_to_v(value, unit)
        elif key in mv:
            dict[key] = voltage_to_mv(value, unit)
        elif key in ns:
            dict[key] = time_to_ns(value, unit)
        elif key in us:  
            dict[key] = time_to_us(value, unit)
        elif key in khz:
            dict[key] = frequency_to_khz(value, unit)
        
def main():
    misslist = []
    root, path = Get_Path() # root = path of project result
    os.chdir(root)
    Create_New_Excel(root)
    wb = opxl.load_workbook(os.path.join(root,'Project_Signal-Integrity(I2C_P)_EVT_Test_Report_Rev.A0.xlsx'))
    ws = wb.active
    ExcelFormat(wb)
    row = 52
    row = VDDVol(ws, row)
    dirs = []
    I2C = ListSubFolders(path)
    for i in I2C: # BMC COM
        dirs2 = ListSubFolders(i)
        for k in dirs2: # I2C2
            dirs.append( "&" + i.split("\\")[-1] + "_" + k.split("\\")[-1] )
            dirs.extend( ListSubFolders(k) )

    for dir in dirs:
        if ( dir[ 0 ] == "&" ):
            row = Header(ws, row, dir[1:len(dir)])
            continue
        print("\n----------------------", dir, "--------------------------------")
        dirname = dir.split("\\")[-1].split("_", 1)[1]
        probname = dirname.split("_", 2)[-1]
        row = DirHeader(ws, row, dirname)
        CLKdict = {}
        DATAdict= {}
        files = [f for f in os.listdir(dir) if f.endswith(".png")]
        for file in files:
            OCR(f'{dir}\\{file}', CLKdict, DATAdict)
            print("CLK: ", CLKdict)
            print("DATA: ", DATAdict)
            print()


        # CLKdict = OCR(r"C:\Users\sam.lu\Desktop\I2C2\04_BMC_to_U117\DATA.jpg")
        # DATAdict = OCR(r"C:\Users\sam.lu\Desktop\I2C2\04_BMC_to_U117\DATA.jpg")

        UnitTransform(CLKdict)
        UnitTransform(DATAdict)

        clk_max = CLKdict["Maximum"] if "Maximum" in CLKdict else None
        data_max = DATAdict["Maximum"] if "Maximum" in DATAdict else None
        clk_min = CLKdict["Minimum"] if "Minimum" in CLKdict else  None
        data_min = DATAdict["Minimum"] if "Minimum" in DATAdict else  None
        row = VolatageHeader(ws, row)
        row1 = row
        row, locname1, locprobe1, misslist = Result_1(ws, row, "CLK", f'PinXX of {probname}', clk_max, "0.7*VDD ~ VDD+0.5", clk_min, "-0.6 ~ 0.3*VDD", "H", [2, 0, 2, 0, 0], misslist, True)
        row, locname2, locprobe2, misslist = Result_1(ws, row, "DATA", f'PinYY of {probname}', data_max, "0.7*VDD ~ VDD+0.5", data_min, "-0.6 ~ 0.3*VDD", "H", [2, 0, 2, 0, 0], misslist,True)   
        row = NumOfTests(row1, row, ws)

        clk_rise = CLKdict["Rise Time"] if "Rise Time" in CLKdict else None 
        data_rise = DATAdict["Rise Time"] if "Rise Time" in DATAdict else None
        clk_fall = CLKdict["Fall Time"] if "Fall Time" in CLKdict else None
        data_fall = DATAdict["Fall Time"] if "Fall Time" in DATAdict else None
        row1 = row
        row = RiseFallTimeHeader(ws, row)
        row, misslist = Result_1(ws, row, locname1, locprobe1, clk_rise, "1000", clk_fall, "300", "H", [2, 0, 2, 0, 0], misslist)
        row, misslist = Result_1(ws, row, locname2, locprobe2, data_rise, "1000", data_fall, "300", "H", [2, 0, 2, 0, 0], misslist)   
        row = NumOfTests(row1, row, ws)

        clk_freq = CLKdict["Frequency"] if "Frequency" in CLKdict else None 
        clk_pos = CLKdict["High Time"] if "High Time" in CLKdict else None 
        clk_neg = CLKdict["Low Time"] if "Low Time" in CLKdict else None                    
        row1 = row
        row = CLKQualityHeader(ws, row)
        row, misslist = Result_2(ws, row, locname1, locprobe1, clk_freq, "0~100", clk_pos, "4", clk_neg, "4.7", "J", [3, 0, 3, 0, 0], misslist)
        row = NumOfTests(row1, row, ws)

        clk_delay = CLKdict["Hold Time"] if "Hold Time" in CLKdict else None
        data_delay = DATAdict["Setup Time"] if "Setup Time" in DATAdict else None
        row1 = row
        row = SetupHoldTimmingHeader(ws, row)
        row, misslist = Result_1(ws, row, locname2, locprobe2, data_delay, "250", clk_delay, "0", "H", [2, 0, 2, 0, 0], misslist)   
        row = NumOfTests(row1, row, ws)        

    wb.save(os.path.join(root,'Project_Signal-Integrity(I2C_P)_EVT_Test_Report_Rev.A0.xlsx'))

    print("total error cell: ", len(misslist))
    print(misslist)

    f = open(os.path.join(root,'fail_to_rec(I2C_P).txt'), 'w')
    for miss in misslist:
        print(miss, file = f)
    f.close()
    print("============ DONE ============")

if __name__ == '__main__':
    main()